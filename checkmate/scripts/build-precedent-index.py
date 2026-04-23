#!/usr/bin/env python3
"""
Build Checkmate's precedent corpus from the `Spare General` Drive folder.

Walks the folder recursively, finds every completed RFP response matrix
(xlsx or csv), parses each matrix, extracts every answered row, and
writes a single precedents.jsonl file suitable for the precedent-search
MCP server.

Usage
-----
    python build-precedent-index.py \
        --credentials ~/.checkmate/credentials.json \
        --token       ~/.checkmate/token.json \
        --folder-name "Spare General" \
        --output      ../data/precedents.jsonl

First-run auth:
    1. Create a Google Cloud project and enable the Drive and Sheets APIs.
    2. Create OAuth client credentials (Desktop app) and download as
       credentials.json. Store it at ~/.checkmate/credentials.json.
    3. Run this script. A browser window opens for OAuth consent; the
       script writes a refresh token to ~/.checkmate/token.json and reuses
       it on subsequent runs.

Output schema (one JSON object per line):
    {
        "id":           "<file_id>:<sheet>:<row>",
        "source_file":  "Laramie RFP - Spare EAM Response Matrix (Updated)",
        "source_row":   "Mandatory Requirements row 3.1.3",
        "agency":       "City of Laramie",
        "year":         2026,
        "requirement":  "Enable preventive maintenance scheduling by time...",
        "verdict":      "Y",
        "comment":      "PM schedules can be set by calendar interval...",
        "url":          "https://docs.google.com/spreadsheets/d/..."
    }
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
except ImportError:
    print(
        "Missing Google API client libraries. Install with:\n"
        "    pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Missing openpyxl. Install with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
MATRIX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.google-apps.spreadsheet",
    "text/csv",
}

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("build-precedent-index")


# ---------------------------------------------------------------------------
# Drive auth + file walking
# ---------------------------------------------------------------------------


def get_drive_service(creds_path: Path, token_path: Path):
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not creds_path.exists():
                raise FileNotFoundError(
                    f"OAuth client credentials not found at {creds_path}. "
                    "Download from Google Cloud Console and place there."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.parent.mkdir(parents=True, exist_ok=True)
        token_path.write_text(creds.to_json())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def find_folder_id(service, folder_name: str) -> str:
    q = (
        f"name = '{folder_name}' and "
        "mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    )
    res = service.files().list(q=q, fields="files(id, name)", pageSize=10).execute()
    items = res.get("files", [])
    if not items:
        raise RuntimeError(f"Drive folder named '{folder_name}' not found")
    if len(items) > 1:
        log.warning(
            "Multiple folders named '%s' found; using the first (%s)",
            folder_name,
            items[0]["id"],
        )
    return items[0]["id"]


def walk_folder(service, folder_id: str) -> Iterable[dict]:
    """Yield file metadata for every matrix-like file below folder_id."""
    pending = [folder_id]
    seen: set[str] = set()
    while pending:
        fid = pending.pop()
        if fid in seen:
            continue
        seen.add(fid)
        page_token = None
        while True:
            res = (
                service.files()
                .list(
                    q=f"'{fid}' in parents and trashed = false",
                    fields=(
                        "nextPageToken,"
                        " files(id, name, mimeType, webViewLink,"
                        " modifiedTime, createdTime, parents)"
                    ),
                    pageSize=100,
                    pageToken=page_token,
                )
                .execute()
            )
            for f in res.get("files", []):
                if f["mimeType"] == "application/vnd.google-apps.folder":
                    pending.append(f["id"])
                elif f["mimeType"] in MATRIX_MIME_TYPES:
                    yield f
            page_token = res.get("nextPageToken")
            if not page_token:
                break


def download_file(service, file: dict) -> bytes:
    fid = file["id"]
    mt = file["mimeType"]
    if mt == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=fid,
            mimeType=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
    else:
        request = service.files().get_media(fileId=fid)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Matrix parsing
# ---------------------------------------------------------------------------


REQUIREMENT_HEADER_PATTERNS = [
    "system requirement",
    "requirement",
    "req #",
    "question",
]
VERDICT_HEADER_PATTERNS = [
    "system meets requirement",
    "bidder response",
    "response",
    "y/n",
    "y/n/p",
    "yes/no",
    "compliance",
]
COMMENT_HEADER_PATTERNS = [
    "bidder details to support response",
    "spare comments",
    "spare response",
    "comments",
    "comment",
    "details",
    "response",
]


@dataclass
class MatrixRow:
    sheet: str
    row_num: int
    requirement: str
    verdict: str
    comment: str


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _find_header_col(row_values: list, patterns: list[str]) -> int | None:
    """Return the earliest column index whose lowered text matches any pattern."""
    best = None
    for idx, val in enumerate(row_values):
        t = _norm(val)
        if not t:
            continue
        for pat in patterns:
            if pat in t:
                return idx
    return best


def _detect_schema(rows: list[list]) -> tuple[int, int, int | None, int | None] | None:
    """
    Return (header_row_idx, req_col, verdict_col, comment_col) or None.
    Scans the first 30 rows for a header row.
    """
    for i, row in enumerate(rows[:30]):
        req_col = _find_header_col(row, REQUIREMENT_HEADER_PATTERNS)
        if req_col is None:
            continue
        verdict_col = _find_header_col(row, VERDICT_HEADER_PATTERNS)
        comment_col = _find_header_col(row, COMMENT_HEADER_PATTERNS)
        if req_col is not None:
            return i, req_col, verdict_col, comment_col
    return None


def parse_xlsx_bytes(data: bytes) -> list[tuple[str, list[list]]]:
    """Return [(sheet_name, rows)] for every sheet in the workbook."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
        out.append((name, rows))
    return out


def parse_csv_bytes(data: bytes) -> list[tuple[str, list[list]]]:
    import csv

    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [list(r) for r in reader]
    return [("Sheet1", rows)]


def extract_rows(file: dict, data: bytes) -> list[MatrixRow]:
    mt = file["mimeType"]
    if mt == "text/csv":
        sheets = parse_csv_bytes(data)
    else:
        sheets = parse_xlsx_bytes(data)
    extracted: list[MatrixRow] = []
    for sheet_name, rows in sheets:
        schema = _detect_schema(rows)
        if not schema:
            continue
        header_idx, req_col, verdict_col, comment_col = schema
        for r_idx in range(header_idx + 1, len(rows)):
            row = rows[r_idx]
            if req_col >= len(row):
                continue
            requirement = str(row[req_col] or "").strip()
            if not requirement or len(requirement) < 8:
                continue
            verdict = ""
            if verdict_col is not None and verdict_col < len(row):
                verdict = str(row[verdict_col] or "").strip()
            comment = ""
            if comment_col is not None and comment_col < len(row):
                comment = str(row[comment_col] or "").strip()
            # Skip rows that are section headers (requirement text but no
            # verdict AND no comment)
            if not verdict and not comment:
                continue
            extracted.append(
                MatrixRow(
                    sheet=sheet_name,
                    row_num=r_idx + 1,
                    requirement=requirement,
                    verdict=verdict,
                    comment=comment,
                )
            )
    return extracted


# ---------------------------------------------------------------------------
# Inference helpers
# ---------------------------------------------------------------------------

AGENCY_RE = re.compile(
    r"(BC Transit|Laramie|MTD|Calgary|NFTA|TCAT|SolTrans|Valley Transit|"
    r"HOLON|GRT|PDRTA|Snyderville|Gulf Coast|GCTD|StarTran|Winnipeg|Oakville|"
    r"Hamilton Street Railway|RTC Washoe|People's Transit|Citibus|Duluth|"
    r"MobilityPLUS|CBRM|Champaign[- ]?Urbana)",
    re.IGNORECASE,
)

YEAR_RE = re.compile(r"(20\d{2})")


def infer_agency(file_name: str) -> str:
    m = AGENCY_RE.search(file_name)
    return m.group(1) if m else ""


def infer_year(file: dict) -> int | None:
    # Prefer explicit year in filename; fall back to modifiedTime year
    m = YEAR_RE.search(file["name"])
    if m:
        return int(m.group(1))
    mt = file.get("modifiedTime") or file.get("createdTime")
    if mt:
        m = YEAR_RE.search(mt)
        if m:
            return int(m.group(1))
    return None


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------


def build_corpus(
    service, folder_id: str, output_path: Path
) -> tuple[int, int]:
    """Returns (files_processed, rows_written)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    files_processed = 0
    rows_written = 0
    with output_path.open("w", encoding="utf-8") as out:
        for file in walk_folder(service, folder_id):
            try:
                data = download_file(service, file)
                matrix_rows = extract_rows(file, data)
            except Exception as e:
                log.warning("skipping %s (%s): %s", file["name"], file["id"], e)
                continue
            if not matrix_rows:
                continue
            files_processed += 1
            agency = infer_agency(file["name"])
            year = infer_year(file)
            url = file.get("webViewLink")
            for mr in matrix_rows:
                out.write(
                    json.dumps(
                        {
                            "id": f"{file['id']}:{mr.sheet}:{mr.row_num}",
                            "source_file": file["name"],
                            "source_row": f"{mr.sheet} row {mr.row_num}",
                            "agency": agency,
                            "year": year,
                            "requirement": mr.requirement,
                            "verdict": mr.verdict,
                            "comment": mr.comment,
                            "url": url,
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
                rows_written += 1
            log.info(
                "indexed %s (%d rows)",
                file["name"],
                len(matrix_rows),
            )
    return files_processed, rows_written


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--credentials",
        type=Path,
        default=Path.home() / ".checkmate" / "credentials.json",
        help="Google OAuth client secrets file",
    )
    p.add_argument(
        "--token",
        type=Path,
        default=Path.home() / ".checkmate" / "token.json",
        help="Cached OAuth refresh token",
    )
    p.add_argument(
        "--folder-name",
        default="Spare General",
        help="Name of the Drive folder to walk",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parent.parent / "data" / "precedents.jsonl",
        help="Output precedents.jsonl path",
    )
    args = p.parse_args()

    log.info("authenticating to Google Drive")
    service = get_drive_service(args.credentials, args.token)

    log.info("locating folder '%s'", args.folder_name)
    folder_id = find_folder_id(service, args.folder_name)

    log.info("building corpus to %s", args.output)
    files, rows = build_corpus(service, folder_id, args.output)

    log.info("done: %d files, %d precedent rows", files, rows)
    if rows == 0:
        log.error(
            "no precedents extracted. check that Spare General contains "
            "completed matrix files with recognizable header rows."
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
