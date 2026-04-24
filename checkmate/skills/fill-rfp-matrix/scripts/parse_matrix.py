#!/usr/bin/env python3
"""
Parse a single matrix file (xlsx or csv) that was downloaded from Drive
into a list of precedent rows. Invoked by the fill-rfp-matrix skill
during its in-session rebuild step for every matrix file discovered in
Spare General.

Usage:
    python parse_matrix.py \
        --input <path to local .xlsx or .csv> \
        --source-file "<original Drive filename>" \
        --source-id   "<Drive file id>" \
        --url         "<Drive webViewLink>" \
        --agency      "<optional agency override>" \
        --year        "<optional year override>" \
        --output      <path to append precedents.jsonl>

If agency and year are omitted they are inferred from the source
filename.

Exit codes:
    0  some rows extracted, written to output
    2  no rows extracted (matrix unrecognized or empty), non-fatal
    3  unrecoverable parse error
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path


REQUIREMENT_HEADER_PATTERNS = [
    "system requirement",
    "requirement",
    "req #",
    "question",
]
VERDICT_HEADER_PATTERNS = [
    "system meets requirement",
    "bidder response",
    "response",
    "y/n",
    "y/n/p",
    "yes/no",
    "compliance",
]
COMMENT_HEADER_PATTERNS = [
    "bidder details to support response",
    "spare comments",
    "spare response",
    "comments",
    "comment",
    "details",
    "response",
]

# Columns whose headers contain any of these substrings are internal scratch
# columns (Checkmate's own "Internal Confidence" and "Internal Reasoning /
# Sources (strip before submit)" columns, SE working notes, etc.) and must
# never be indexed as precedent answer content. Added in v1.3.6 after observing
# internal-note text leaking from the corpus into customer-facing drafts.
INTERNAL_COLUMN_PATTERNS = [
    "internal confidence",
    "internal reasoning",
    "internal sources",
    "internal notes",
    "strip before submit",
    "do not submit",
    "do not send",
    "do not share",
    "confidential",
    "se notes",
    "se working",
    "spare internal",
    "for internal use",
]

# Row-level content markers indicating a comment is SE scratch-note text rather
# than a polished customer-facing answer. Any row whose comment column contains
# one of these gets dropped on index. Patterns are deliberately narrow to avoid
# false positives on legitimate answers (e.g. "via a custom field" is fine;
# "may need custom" as a standalone hedge is the scratch-note signal).
#
# The primary defense against internal-content leakage is INTERNAL_COLUMN_PATTERNS
# above (never index from an internal column in the first place). This list is
# the backup for cases where internal notes ended up in the actual comment
# column.
INTERNAL_CONTENT_MARKERS = [
    # Literal TBD / pending markers (very rarely appear in polished answers)
    r"\btbd\b",
    r"\btbd[-:]",
    # High-specificity internal-decision hedging (phrases that a polished
    # customer-facing answer essentially never contains)
    r"\bmight not want\b",
    r"\bmight have complexity\b",
    r"\bnot likely to have\b",
    r"\brequires decision on pursuit\b",
    # Obvious typos catching scratch-note drift (the specific misspelling
    # observed in a live v1.3.1 run)
    r"\busse\b",
]
INTERNAL_CONTENT_RE = re.compile("|".join(INTERNAL_CONTENT_MARKERS), re.IGNORECASE)

AGENCY_RE = re.compile(
    r"(BC Transit|Laramie|MTD|Calgary|NFTA|TCAT|SolTrans|Valley Transit|"
    r"HOLON|GRT|PDRTA|Snyderville|Gulf Coast|GCTD|StarTran|Winnipeg|Oakville|"
    r"Hamilton Street Railway|RTC Washoe|People's Transit|Citibus|Duluth|"
    r"MobilityPLUS|CBRM|Champaign[- ]?Urbana)",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"(20\d{2})")


@dataclass
class MatrixRow:
    sheet: str
    row_num: int
    requirement: str
    verdict: str
    comment: str


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _is_internal_column(header: str) -> bool:
    """Return True if the column header marks an internal / non-customer-facing
    column that must never be indexed as precedent content."""
    t = _norm(header)
    if not t:
        return False
    return any(pat in t for pat in INTERNAL_COLUMN_PATTERNS)


def _find_header_col(row_values, patterns):
    """Return the earliest column index whose header matches `patterns` AND is
    NOT an internal / strip-before-submit column."""
    for idx, val in enumerate(row_values):
        t = _norm(val)
        if not t:
            continue
        if _is_internal_column(val):
            # Never match internal columns even if they otherwise match
            # e.g. an "Internal Reasoning / Sources" column should not be
            # mistaken for the comment column.
            continue
        for pat in patterns:
            if pat in t:
                return idx
    return None


def _detect_schema(rows):
    for i, row in enumerate(rows[:30]):
        req_col = _find_header_col(row, REQUIREMENT_HEADER_PATTERNS)
        if req_col is None:
            continue
        verdict_col = _find_header_col(row, VERDICT_HEADER_PATTERNS)
        comment_col = _find_header_col(row, COMMENT_HEADER_PATTERNS)
        return i, req_col, verdict_col, comment_col
    return None


def _parse_xlsx(path: Path):
    """Returns [(sheet_name, rows)]. Requires openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "openpyxl is not available in this environment. The "
            "fill-rfp-matrix skill should install it at session "
            "start via `pip install openpyxl --break-system-packages`.",
            file=sys.stderr,
        )
        sys.exit(3)
    wb = load_workbook(str(path), data_only=True, read_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        out.append((name, rows))
    return out


def _parse_csv(path: Path):
    text = path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [("Sheet1", [list(r) for r in reader])]


def extract_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheets = _parse_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        sheets = _parse_xlsx(path)
    else:
        raise ValueError(f"unsupported file extension: {suffix}")

    extracted: list[MatrixRow] = []
    for sheet_name, rows in sheets:
        schema = _detect_schema(rows)
        if not schema:
            continue
        header_idx, req_col, verdict_col, comment_col = schema
        for r_idx in range(header_idx + 1, len(rows)):
            row = rows[r_idx]
            if req_col >= len(row):
                continue
            requirement = str(row[req_col] or "").strip()
            if not requirement or len(requirement) < 8:
                continue
            verdict = ""
            if verdict_col is not None and verdict_col < len(row):
                verdict = str(row[verdict_col] or "").strip()
            comment = ""
            if comment_col is not None and comment_col < len(row):
                comment = str(row[comment_col] or "").strip()
            # Skip section-header rows (requirement text, no verdict, no comment)
            if not verdict and not comment:
                continue
            # Skip rows whose comment is clearly SE scratch-note text. These
            # are draft / working matrices where internal notes ended up in
            # the comment column instead of a polished customer answer. See
            # v1.3.5 Rule 26 for the failure mode and INTERNAL_CONTENT_MARKERS
            # above for the scan patterns.
            if comment and INTERNAL_CONTENT_RE.search(comment):
                continue
            extracted.append(
                MatrixRow(
                    sheet=sheet_name,
                    row_num=r_idx + 1,
                    requirement=requirement,
                    verdict=verdict,
                    comment=comment,
                )
            )
    return extracted


def infer_agency(file_name: str) -> str:
    m = AGENCY_RE.search(file_name)
    return m.group(1) if m else ""


def infer_year(file_name: str) -> int | None:
    m = YEAR_RE.search(file_name)
    return int(m.group(1)) if m else None


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", type=Path, required=True)
    p.add_argument("--source-file", required=True)
    p.add_argument("--source-id", required=True)
    p.add_argument("--url", default="")
    p.add_argument("--agency", default="")
    p.add_argument("--year", default="")
    p.add_argument("--output", type=Path, required=True)
    args = p.parse_args()

    try:
        rows = extract_rows(args.input)
    except Exception as e:  # noqa: BLE001
        print(f"parse error in {args.source_file}: {e}", file=sys.stderr)
        return 3

    if not rows:
        print(f"no rows extracted from {args.source_file}", file=sys.stderr)
        return 2

    agency = args.agency or infer_agency(args.source_file)
    year_str = args.year or ""
    year = int(year_str) if year_str.isdigit() else infer_year(args.source_file)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("a", encoding="utf-8") as out:
        for r in rows:
            out.write(
                json.dumps(
                    {
                        "id": f"{args.source_id}:{r.sheet}:{r.row_num}",
                        "source_id": args.source_id,
                        "source_file": args.source_file,
                        "source_row": f"{r.sheet} row {r.row_num}",
                        "agency": agency,
                        "year": year,
                        "requirement": r.requirement,
                        "verdict": r.verdict,
                        "comment": r.comment,
                        "url": args.url or None,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

    print(f"indexed {args.source_file}: {len(rows)} rows", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
